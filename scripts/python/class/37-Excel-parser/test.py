from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill
import dump_cell

wb = load_workbook('revenue.xlsx')

print wb.get_sheet_names()
for wsheet in wb.get_sheet_names():
    print wsheet

wsheet = wb.get_sheet_by_name('sales')

print wsheet[2][2].value
print wsheet[3][2].value

cell = wsheet[2][2]
print cell.value

rgb=[255,0,0]
color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])

#cell.fill = PatternFill(fill_type="solid", start_color=colors.GREEN, end_color=colors.RED)
cell.fill = PatternFill(fill_type="solid", start_color=colors.GREEN)

#cell.fill = PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
#PatternFill(bgColor="000000")

wb.save('revenue.xlsx')
#exit(1)
