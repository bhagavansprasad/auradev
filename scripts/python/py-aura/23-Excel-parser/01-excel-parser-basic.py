import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill


def print_all_worksheet_names(wbname):
	wb = load_workbook(wbname)

	#print worksheet names
	print(wb.get_sheet_names())

	for wsheet in wb.get_sheet_names():
		print(wsheet)
	print ("")

def print_cell_values(wbname):
	wb = load_workbook(wbname)

	#opening worksheet in workbook
	wsheet = wb.get_sheet_by_name('sales')

	#max rows and cols
	rcount =  wsheet.max_row
	ccount =  wsheet.max_column
	print("max rows filled :", rcount)
	print("max cols filled :", ccount)
	
	#print cell value
	print("A2           :", wsheet['A2'].value) 
	print("A3           :", wsheet['C3'].value) 

	#print in row and col format
	print("wsheet[2][0] :", wsheet[2][0].value) 
	print("wsheet[3][2] :", wsheet[3][2].value) 
	print ("")
	return

def print_rows_columns1(wbname):
	wb = load_workbook(wbname)

	#opening worksheet in workbook
	wsheet = wb.get_sheet_by_name('sales')

	#max rows and cols
	rcount =  wsheet.max_row
	ccount =  wsheet.max_column

	#print rows and columns
	for row in range(1, rcount+1):
		flag = 1
		for col in range (0, ccount):
			if (wsheet[row][col].value != None):
				#print (str(wsheet[row][col].value)),
				print((wsheet[row][col].value), end=' ')
				flag = 0

		if (flag == 1):
			break;

		print("")

	print ("")

def print_rows_columns2(wbname):

	#opening workbook
	wb = load_workbook(wbname)

	#opening worksheet in workbook
	wsheet = wb.get_sheet_by_name('sales')

	#max rows and cols
	rcount =  wsheet.max_row
	ccount =  wsheet.max_column

	#print rows and columns
	for row in range(1, rcount+1):
		if not any(cell.value for cell in wsheet[row]):
			print("The %d row is empty" % row)
			break

		for col in range (0, ccount):
			if (wsheet[row][col].value != None):
				print((str(wsheet[row][col].value)), end=' ')

		print ("")


def highlight_cell_by_value(wbname, wsheet_name, cell_value):

	#opening workbook
	wb = load_workbook(wbname)

	#opening worksheet in workbook
	wsheet = wb.get_sheet_by_name(wsheet_name)

	#max rows and cols
	rcount =  wsheet.max_row
	ccount =  wsheet.max_column

	#Highlight cell, if the cell value is Vinay
	#Change cell font size to 12, color Red, bold

	#print (type(wsheet[1]))
	#print (wsheet[1])

	for row in range(1, rcount+1):
		if not any(cell.value for cell in wsheet[row]):
			print("The %d row is empty" % row)
			break

		for col in range (0, ccount):
			if (wsheet[row][col].value != None and wsheet[row][col].value == cell_value ):
				print(wsheet[row][col].font)
				wsheet[row][col].font = Font(size=15, bold=True, color=colors.GREEN)
				print("")

	wb.save(wbname)

#Highlight row background if the cell value is as cell_value
#Change row background to yellow
def change_background_cell_by_value(wbname, wsheet_name, cell_value):
	#opening workbook
	wb = load_workbook(wbname)

	#opening worksheet in workbook
	wsheet = wb.get_sheet_by_name(wsheet_name)

	#max rows and cols
	rcount =  wsheet.max_row
	ccount =  wsheet.max_column

	for row in range(1, rcount):
		if not any(cell.value for cell in wsheet[row]):
			print("The %d row is empty" % row)
			break

		for col in range (0, ccount):
			if (wsheet[row][col].value != None and wsheet[row][col].value == cell_value ):
				print(wsheet[row][col].fill)
				print(str(wsheet[row][col].value))

				#for cell in wsheet[row:row]:
				for cell in wsheet[row]:
					cell.fill = PatternFill(fill_type="solid", start_color=colors.YELLOW)

	wb.save(wbname)

def main():
	#print_all_worksheet_names('shared/revenue.xlsx')
	#print_cell_values('shared/revenue.xlsx')
	#print_rows_columns1('shared/revenue.xlsx')
	#print_rows_columns2('shared/revenue.xlsx')
	#highlight_cell_by_value('shared/revenue.xlsx', "sales", "Ashish")
	#change_background_cell_by_value('shared/revenue.xlsx', "sales", "Kavitha")

if (__name__ == '__main__'):
	main()


'''
#Highlight cell, if the cell value is Vinay
#Change cell font size to 12, color Red, bold
for row in range(1, rcount+1):
    if not any(cell.value for cell in wsheet[row]):
        print("The %d row is empty" % row)
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Vinay" ):
            print((wsheet[row][col].font))
            wsheet[row][col].font = Font(size=12, bold=True, color=colors.RED)
            print("")

#Highlight row if the cell value is Ashish
#Change row font size to 12, color GREEN, bold
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print("The %d row is empty" % row)
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Ashish" ):
            for cell in wsheet[row:row]:
                cell.font = Font(size=12, bold=True, color=colors.GREEN)

#Highlight column background in case cell value is > 70000
for row in range(1, rcount):
    if (wsheet[row][4].value != None and int(wsheet[row][4].value) > int(70000) ):
        print((str(wsheet[row][4].value)))
        cell = wsheet[row][4]
        cell.fill = PatternFill(fill_type="solid", start_color=colors.GREEN)

for word in dir(wsheet['C3']):
    print("print \"%-20s :\", cell.%s" %  (word, word))

wb.save('shared/revenue.xlsx')
#dump_cell.dump_cell_properties(wsheet[1][0])
'''

exit(1)
