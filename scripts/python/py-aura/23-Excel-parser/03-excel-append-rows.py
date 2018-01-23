from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill, Protection

#print worksheet names
#print wb.get_sheet_names()

def list_worksheets_by_wbook(wbook):
    wb = load_workbook(wbook)
    print("worksheet names...")
    for wsheet in wb.get_sheet_names():
        print(("'%s'" % (wsheet)), end=' ')
    print("")

def append_rows_by_name(wbook, dst_wname, src_wname, filter_str):
    wb = load_workbook(wbook)

    #opening source worksheet in workbook
    swsheet = wb.get_sheet_by_name(src_wname)

    #Create destination worksheet, if does not exists
    try:
        dwsheet = wb.get_sheet_by_name(dst_wname)
        print("Worksheet '%s' found" % (dst_wname))

    except:
        print("Worksheet '%s' not found" % (dst_wname))
        print("Creating new worksheet :'%s'" % (dst_wname))
        dwsheet = wb.create_sheet(dst_wname, 0)

    #max rows and cols
    srcount =  swsheet.max_row
    sccount =  swsheet.max_column
    print("%s: max row:col (%d:%d)" % (src_wname, srcount, sccount))

    drcount =  dwsheet.max_row
    dccount =  dwsheet.max_column
    print("%s: max row:col (%d:%d)" % (dst_wname, drcount, dccount))

    #copy titles
    #copy titles only if it is new sheet
    if (drcount <= 1):
        row = swsheet[1]
        frow = [cell.value for cell in row]
        dwsheet.append(frow)

    #dump rows
    for row in dwsheet.iter_rows():
        for cell in row:
            print(cell.value, " ", end=' ')
        print("")

    #copy all rows that matched with given string
    for row in swsheet.iter_rows():
        if (row[2].value != None and row[2].value == filter_str):
            frow = [cell.value for cell in row]
            print("appending {0}".format(frow))
            dwsheet.append(frow)

    print("Saving :%s" % (wbook))
    wb.save(wbook)

def main():
    wbook = 'shared/revenue.xlsx'
    src_wname = "sales"
    dst_wname = "filtered"
    filter_str = "Ashish"

    list_worksheets_by_wbook(wbook)
    append_rows_by_name(wbook, dst_wname, src_wname, filter_str)

if (__name__ == '__main__'):
    main()

